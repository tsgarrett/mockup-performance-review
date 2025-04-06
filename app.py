import streamlit as st
import pandas as pd
from datetime import date
import io
from openpyxl.styles import PatternFill

# Recommendations mapping per ad stage
# Note: These are triggered based on the 'Flag Reason' determined by the logic below.
recommendations_by_stage = {
    "Mockup": {
        "Low CTR": "Link Click CTR is below threshold. Test radically different visuals or curiosity-driven headlines. [Ref: Stage 3 – 03:45:10]",
        "High CPC": "CPC above threshold in mockup phase. Consider stronger hooks or clearer visual cues. [Ref: Stage 3 – 03:52:10]",
        "No engagement": "Spent over threshold with minimal/no link clicks. Kill ad and test a fresh creative angle. [Ref: Stage 3 – 03:49:22]",
        "Keep": "Metrics look good for this stage. Consider moving into Cycle 1 if stable. [Ref: Stage 3 – 03:58:00]"
        # "Insufficient Data" will be handled separately
    },
    "Cycle 1": {
        "Low CTR": "Link Click CTR below threshold. Try tightening your hook or using more urgency. [Ref: Stage 4 – 04:20:03]",
        "High CPC": "CPC is above your threshold. Adjust targeting or test broader audiences. [Ref: Stage 4 – 04:31:50]",
        "No engagement": "Spent over threshold with minimal/no link clicks. Kill ad or duplicate with bold creative shift. [Ref: Stage 4 – 04:12:45]",
        "Keep": "Good signals. Let it run and monitor CPC/CTR closely. [Ref: Stage 4 – 04:40:00]"
        # "Insufficient Data" will be handled separately
    },
    "Cycle 2": {
        "Low CTR": "Link Click CTR below threshold for scaling. Consider fatigue or ad set saturation. [Ref: Stage 5 – 05:01:00]",
        "High CPC": "CPC too high for scaling. Restructure ad set or test budget split. [Ref: Stage 5 – 05:06:12]",
        "No ROAS": "Spent over threshold with no ROAS. Review funnel, page load time, and offer clarity. [Ref: Stage 5 – 05:02:17]",
        "Keep": "Strong performer. Consider scaling or cloning into new audience segments. [Ref: Stage 5 – 05:14:03]"
        # "Insufficient Data" will be handled separately
    }
}

# --- Streamlit App Configuration ---
st.set_page_config(page_title="Ad Performance Review", layout="centered")
if "upload_key" not in st.session_state:
    st.session_state.upload_key = 0

st.title("📊 Ad Performance Review")
st.markdown("A stage-aware review tool that adapts recommendations based on your campaign phase.")

# --- ADDED DISCLAIMER ---
st.info(
    """
    **Disclaimer:** This tool is intended as an **aid** to your ad review process and should complement, **not replace**, your own judgment and analysis.
    It was created by a member for members and is **not an official product of, nor endorsed by, the WeScale Group Founders.**
    Always consider the full context of your campaigns when making decisions.
    """,
    icon="ℹ️"
)
# --- END DISCLAIMER ---


# --- User Inputs ---

# Step 0: Choose Ad Stage
ad_stage = st.radio(
    "Select Ad Stage:",
    ["Mockup", "Cycle 1", "Cycle 2"],
    horizontal=True
)

st.subheader("Step 1: Define Evaluation Thresholds")

# Use columns for better layout
col1, col2 = st.columns(2)

with col1:
    # CTR threshold (configurable) - Based on Link Click CTR now
    ctr_threshold_percent = st.number_input(
        "Minimum Acceptable Link Click CTR (%)",
        min_value=0.0,
        max_value=100.0,
        value=0.75, # Default value - Adjust if needed for Link CTR context
        step=0.05,
        format="%.2f"
    )
    # Convert percentage to decimal for comparison
    ctr_threshold = ctr_threshold_percent / 100.0

    # Spend threshold for initial evaluation (Mockup, Cycle 1)
    initial_spend_threshold = st.number_input(
        "Spend threshold for Mockup/Cycle 1 eval ($)",
        min_value=0.0,
        value=5.0, # Default value
        step=0.50,
        format="%.2f"
    )

with col2:
    # CPC threshold selector
    cpc_threshold = st.selectbox(
        "Maximum Acceptable CPC ($)",
        [0.50, 0.75, 1.00, 1.25, 1.50, 1.75, 2.00, 2.50, 3.00, 5.00], # Expanded options
        index=2 # Default to 1.00
    )

    # Spend threshold for Cycle 2 ROAS evaluation
    cycle2_spend_threshold = st.number_input(
        "Spend threshold for Cycle 2 eval ($)",
        min_value=0.0,
        value=15.0, # Default value
        step=1.00,
        format="%.2f"
    )

st.subheader("Step 2: Upload Your File")
# Updated note about required columns and CTR type
st.markdown(
    """
    **Required columns:** `Ad name`, `Amount spent (USD)`, `CTR (link click-through rate)`,
    `CPC (cost per link click) (USD)`, `Link clicks`, `Purchase ROAS (return on ad spend)`.

    *Note: This script uses **CTR (link click-through rate)** for analysis. Ensure your Facebook export includes this exact column name.*
    """
)
uploaded_file = st.file_uploader(
    "Upload your Facebook Ads Excel export (.xlsx)",
    type=["xlsx"],
    key=st.session_state.upload_key
)

# --- Main Processing Logic ---
if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        # UPDATED required_cols list
        required_cols = [
            "Ad name", "Amount spent (USD)", "CTR (link click-through rate)", # Changed from CTR (all)
            "CPC (cost per link click) (USD)", "Link clicks",
            "Purchase ROAS (return on ad spend)"
        ]
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            st.error(f"Missing required columns: {', '.join(missing)}")
            st.stop()

        # Ensure numeric types where necessary, coerce errors to NaN
        # UPDATED to include the correct CTR column name
        for col in ["Amount spent (USD)", "CTR (link click-through rate)", "CPC (cost per link click) (USD)", "Link clicks", "Purchase ROAS (return on ad spend)"]:
             # Add check if column exists before converting, although 'missing' check should cover this
            if col in df.columns:
                 df[col] = pd.to_numeric(df[col], errors='coerce')
            # No else needed here because the 'missing' check above would have stopped execution

        # --- Evaluation Function (Corrected Logic & using Link Click CTR) ---
        def evaluate(row):
            spend = row["Amount spent (USD)"]
            # Use the configured CTR threshold (decimal)
            # UPDATED to use 'CTR (link click-through rate)'
            ctr = row["CTR (link click-through rate)"]
            # Use the configured CPC threshold
            cpc = row["CPC (cost per link click) (USD)"]
            clicks = row["Link clicks"] # Using 'Link clicks' count directly is robust for engagement check
            roas = row["Purchase ROAS (return on ad spend)"]

            # Minimum number of link clicks to consider engagement 'present'
            # If Link clicks is 0 or NaN after spending threshold, flag as no engagement.
            min_clicks_for_engagement = 1

            if ad_stage == "Mockup":
                # Check 1: Has it spent enough to evaluate?
                if spend < initial_spend_threshold:
                    return "N", "Insufficient Data" # Not enough spend yet

                # Check 2: Spent threshold, but no engagement (based on Link Clicks)?
                if pd.isna(clicks) or clicks < min_clicks_for_engagement:
                    return "Y", "No engagement"

                # Check 3: Low Link Click CTR?
                # Check if Link Click CTR is NaN (unlikely if clicks >= 1, but good practice) or below threshold
                if pd.isna(ctr) or ctr < ctr_threshold:
                    return "Y", "Low CTR"

                # Check 4: High CPC?
                # Check if CPC is NaN or above threshold
                if pd.isna(cpc) or cpc > cpc_threshold:
                    return "Y", "High CPC"

                # If none of the above kill criteria are met
                return "N", "Keep"

            elif ad_stage == "Cycle 1":
                # Check 1: Has it spent enough to evaluate?
                if spend < initial_spend_threshold:
                    return "N", "Insufficient Data" # Not enough spend yet

                # Check 2: Spent threshold, but no engagement (based on Link Clicks)?
                if pd.isna(clicks) or clicks < min_clicks_for_engagement:
                    return "Y", "No engagement"

                # Check 3: Low Link Click CTR?
                if pd.isna(ctr) or ctr < ctr_threshold:
                    return "Y", "Low CTR"

                # Check 4: High CPC?
                if pd.isna(cpc) or cpc > cpc_threshold:
                     return "Y", "High CPC"

                # If none of the above kill criteria are met
                return "N", "Keep"

            elif ad_stage == "Cycle 2":
                # Check 1: Has it spent enough for Cycle 2 evaluation?
                if spend < cycle2_spend_threshold:
                    return "N", "Insufficient Data" # Not enough spend yet

                # Check 2: No ROAS after significant spend? (Prioritize this in Cycle 2)
                if pd.isna(roas) or roas <= 0:
                    return "Y", "No ROAS"

                # Check 3: Low Link Click CTR? (Still relevant for scaling health)
                if pd.isna(ctr) or ctr < ctr_threshold:
                    return "Y", "Low CTR"

                # Check 4: High CPC? (Check after ROAS and CTR)
                if pd.isna(cpc) or cpc > cpc_threshold:
                    return "Y", "High CPC"

                # If none of the above kill criteria are met
                return "N", "Keep"

            # Fallback (should not be reached if logic is correct)
            return "N", "Review Manually"

        # Apply the evaluation function
        df[["Kill Criteria Met? (Y/N)", "Flag Reason"]] = df.apply(evaluate, axis=1, result_type="expand")

        # Define Action mapping (more nuanced)
        action_mapping = {
            "Keep": "Keep Running",
            "Insufficient Data": "Keep Running (Monitor)",
            "Low CTR": "Optimize/Review",
            "High CPC": "Optimize/Review",
            "No engagement": "Pause/Kill",
            "No ROAS": "Pause/Kill",
            "Review Manually": "Review Manually"
        }

        # --- Prepare Output DataFrame ---
        review = pd.DataFrame({
            "Date of Report": [date.today().strftime("%Y-%m-%d")] * len(df), # Format date
            "Ad Name": df["Ad name"],
            "Amount Spent (USD)": df["Amount spent (USD)"].round(2),
            # UPDATED: Display Link Click CTR as percentage, handle NaN
            "Link CTR (%)": df["CTR (link click-through rate)"].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "N/A"),
            # Handle NaN for clicks (using the Link clicks column)
            "Link Clicks": df["Link clicks"].apply(lambda x: int(x) if pd.notna(x) else "N/A"),
            # Round CPC, handle NaN
            "CPC (USD)": df["CPC (cost per link click) (USD)"].apply(lambda x: f"${x:.2f}" if pd.notna(x) else "N/A"),
             # Round ROAS, handle NaN
            "ROAS": df["Purchase ROAS (return on ad spend)"].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A"),
            "Kill Criteria Met? (Y/N)": df["Kill Criteria Met? (Y/N)"],
            "Flag Reason": df["Flag Reason"], # Show the reason determined by evaluate()
            # Apply the nuanced action mapping
            "Action to Take": df["Flag Reason"].map(action_mapping).fillna("Review Manually"),
            # Get recommendation based on stage and flag reason (recommendation text updated slightly for clarity)
            "Detailed Recommendation": df["Flag Reason"].apply(
                lambda r: recommendations_by_stage[ad_stage].get(r, "Check metrics manually based on flag reason.") # Provide fallback
            ),
        })

        # Add empty Notes column for user input in Excel
        review["Notes"] = ""

        # --- Display Results in Streamlit ---
        st.subheader("📊 Summary")
        flagged_count = (review['Kill Criteria Met? (Y/N)'] == 'Y').sum()
        st.write(f"**Total Ads Processed:** {len(review)}")
        st.write(f"**Ads Flagged for Action (Y):** {flagged_count}")

        # Highlighting function for the DataFrame display
        def highlight_rows(row):
            if row['Kill Criteria Met? (Y/N)'] == 'Y':
                color = '#ffe6e6' # Light red
            elif row['Flag Reason'] == 'Insufficient Data':
                color = '#fff9e6' # Light yellow/orange
            else: # Keep or other 'N' flags
                color = '#e6ffe6' # Light green
            return [f'background-color: {color}'] * len(row)

        st.subheader("Step 3: Review Results")
        st.dataframe(review.style.apply(highlight_rows, axis=1), height=400) # Set height for better scrolling

        # --- Download Functionality ---
        st.subheader("Step 4: Download Report")
        buffer = io.BytesIO()
        # Use ExcelWriter to apply formatting
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            review.to_excel(writer, index=False, sheet_name="Ad Review")
            worksheet = writer.sheets["Ad Review"]

            # Define fills for Excel conditional formatting
            red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid") # Light Red
            green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid") # Light Green
            yellow_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid") # Light Yellow/Orange

            # Apply formatting row by row based on 'Kill Criteria Met?' and 'Flag Reason'
            # Find the column index for matching (1-based index for openpyxl)
            kill_col_idx = review.columns.get_loc("Kill Criteria Met? (Y/N)") + 1
            reason_col_idx = review.columns.get_loc("Flag Reason") + 1

            for row_idx in range(2, len(review) + 2): # Start from row 2 (after header)
                kill_val = worksheet.cell(row=row_idx, column=kill_col_idx).value
                reason_val = worksheet.cell(row=row_idx, column=reason_col_idx).value

                fill_to_apply = None
                if kill_val == 'Y':
                    fill_to_apply = red_fill
                elif reason_val == 'Insufficient Data':
                     fill_to_apply = yellow_fill
                else: # 'N' and not insufficient data
                    fill_to_apply = green_fill

                if fill_to_apply:
                    for col_idx in range(1, len(review.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = fill_to_apply

            # Auto-adjust column widths (optional but helpful)
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name like 'A'
                # Adjust max_length check to include header
                header_cell = worksheet[f"{column}1"]
                if header_cell.value:
                     max_length = len(str(header_cell.value))

                for cell in col:
                     # Skip header row as it's already checked
                    if cell.row == 1:
                        continue
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except:
                        pass
                # Add padding to max_length
                adjusted_width = (max_length + 2) * 1.2
                # Set a reasonable max width
                if adjusted_width > 50:
                    adjusted_width = 50
                worksheet.column_dimensions[column].width = adjusted_width


        st.download_button(
            label="📥 Download Formatted Ad Review Sheet (.xlsx)",
            data=buffer.getvalue(),
            file_name=f"Ad_Review_{ad_stage}_{date.today().strftime('%Y%m%d')}.xlsx", # Dynamic filename
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except FileNotFoundError:
        st.error("Uploaded file not found. Please ensure the file was uploaded correctly.")
    except KeyError as e:
         st.error(f"Missing expected column in the uploaded file: {e}. Please ensure your file includes all required columns with the exact names specified.")
         st.info(f"Required columns are: {', '.join(required_cols)}")
    except Exception as e:
        st.error(f"An error occurred during processing: {e}")
        st.exception(e) # Show full traceback for debugging if needed

# --- Start Over Button ---
st.divider()
if st.button("🔄 Start Over / Upload New File"):
    # Clear previous errors or results by forcing a full rerun
    st.session_state.upload_key += 1 # Change key to force re-render of file_uploader
    st.rerun()